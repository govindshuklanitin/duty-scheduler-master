import os
from flask import Flask, render_template, request, send_file, jsonify
from datetime import datetime, timedelta
import calendar
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, Color
from openpyxl.utils import get_column_letter
from functools import lru_cache
import logging
from werkzeug.middleware.profiler import ProfilerMiddleware
from io import BytesIO
from reportlab.lib.pagesizes import landscape, A4
from reportlab.pdfgen import canvas
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer
from reportlab.lib import colors
from reportlab.platypus import Paragraph
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

app = Flask(__name__)

# Enable profiling conditionally
if os.environ.get('ENABLE_PROFILING') == '1':
    app.config['PROFILE'] = True
    app.wsgi_app = ProfilerMiddleware(app.wsgi_app, restrictions=[30])

class DutyScheduler:
    def __init__(self):
        self.shifts = {
            'A': '06:00-14:00',
            'B': '14:00-22:00',
            'C': '22:00-06:00',
            'G': 'General',
            'R': 'Rest'
        }
        self.shift_rotation = {'A': 'C', 'C': 'B', 'B': 'A'}
        # Color mappings for the Excel output
        self.shift_colors = {
            'A': '3b82f6',  # Blue
            'B': '8b5cf6',  # Purple
            'C': 'f97316',  # Orange
            'G': '14b8a6',  # Teal
            'R': '64748b'   # Gray
        }
        
    @staticmethod
    def get_day_name(year, month, day):
        return calendar.day_abbr[calendar.weekday(year, month, day)].upper()

    def get_next_shift(self, current_shift):
        return self.shift_rotation.get(current_shift, current_shift)

    def generate_schedule(self, employees_data, year, month):
        """
        Generate the duty schedule
        
        Args:
            employees_data: List or tuple containing employee data
            year: The year to generate the schedule for
            month: The month to generate the schedule for
            
        Returns:
            A dictionary with the schedule for each employee
        """
        # 1. Debug input information
        logger.info(f"Input employees_data type: {type(employees_data)}")
        logger.info(f"Example of employees_data: {str(employees_data)[:200]}")  # Limit output size
        
        # 2. Ensure we're working with a list
        if not isinstance(employees_data, list):
            employees_data = list(employees_data) if isinstance(employees_data, tuple) else [employees_data]
        
        # 3. Create a clean list of employee dictionaries
        cleaned_employees = []
        
        for emp in employees_data:
            # If we have a tuple, convert it to a dict
            if isinstance(emp, tuple):
                emp_dict = {}
                try:
                    # Handle tuple of key-value tuples
                    for item in emp:
                        if isinstance(item, tuple) and len(item) >= 2:
                            key, value = item[0], item[1]
                            emp_dict[key] = value
                except Exception as e:
                    logger.error(f"Error converting tuple to dict: {str(e)}, tuple: {emp}")
                    continue
            elif isinstance(emp, dict):
                # Already a dict
                emp_dict = emp
            else:
                # Unknown format
                logger.error(f"Unsupported employee data type: {type(emp)}, value: {emp}")
                continue
            
            # Ensure all required keys are present
            required_keys = ['name', 'code', 'post', 'start_shift', 'rest_day']
            if not all(key in emp_dict for key in required_keys):
                missing = [key for key in required_keys if key not in emp_dict]
                logger.error(f"Missing required keys {missing} in employee data: {emp_dict}")
                continue
            
            # Convert rest_day to int if it's not already
            try:
                emp_dict['rest_day'] = int(emp_dict['rest_day'])
            except (ValueError, TypeError):
                logger.error(f"Invalid rest_day value: {emp_dict.get('rest_day')}")
                continue
                
            cleaned_employees.append(emp_dict)
        
        # If we have no valid employees, return an empty schedule
        if not cleaned_employees:
            logger.error("No valid employees after data cleaning")
            return {}
            
        # 4. Generate the schedule
        logger.info(f"Generating schedule with {len(cleaned_employees)} valid employees")
        num_days = calendar.monthrange(year, month)[1]
        schedule = {}
        
        # Pre-calculate weekdays for the entire month
        weekdays = [calendar.weekday(year, month, day) for day in range(1, num_days + 1)]
        
        for emp in cleaned_employees:
            schedule[emp['name']] = {
                'code': emp['code'],
                'post': emp['post'],
                'shifts': []
            }
            
            current_shift = emp['start_shift']
            # Convert Sunday from 0 to 6
            rest_day = 6 if emp['rest_day'] == 0 else emp['rest_day'] - 1
            was_rest_day = False
            
            for day_index, day in enumerate(range(1, num_days + 1)):
                # Check if it's a rest day using pre-calculated weekdays
                if weekdays[day_index] == rest_day:
                    schedule[emp['name']]['shifts'].append('R')
                    was_rest_day = True
                else:
                    if was_rest_day and current_shift != 'G':
                        current_shift = self.get_next_shift(current_shift)
                        was_rest_day = False
                    schedule[emp['name']]['shifts'].append(current_shift)
        
        return schedule

@app.route('/')
def index():
    try:
        logger.info("Rendering index page")
        return render_template('index.html')
    except Exception as e:
        logger.error(f"Error rendering index page: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/generate', methods=['POST'])
def generate():
    try:
        start_time = datetime.now()
        data = request.get_json()
        
        if not data:
            logger.warning("No JSON data received")
            return jsonify({"error": "No data provided"}), 400
            
        year = int(data.get('year', datetime.now().year))
        month = int(data.get('month', datetime.now().month))
        employees_data = data.get('employees', [])
        
        # Add detailed logging for debugging
        logger.info(f"Raw data received: {str(data)[:200]}")
        logger.info(f"Employees data type: {type(employees_data)}")
        
        if not employees_data:
            logger.warning("No employee data received")
            return jsonify({"error": "No employee data provided"}), 400
        
        logger.info(f"Generating schedule for {month}/{year} with {len(employees_data)} employees")
        
        # Try to see what's in the first employee
        if employees_data and len(employees_data) > 0:
            first_emp = employees_data[0]
            logger.info(f"First employee type: {type(first_emp)}")
            logger.info(f"First employee data: {first_emp}")
            
            # If it's a dictionary, check the keys
            if isinstance(first_emp, dict):
                logger.info(f"First employee keys: {first_emp.keys()}")
        
        # Generate the schedule
        scheduler = DutyScheduler()
        schedule = scheduler.generate_schedule(employees_data, year, month)
        
        # Check if we got an empty schedule (indicates error)
        if not schedule:
            logger.error("Generated schedule is empty, likely due to data errors")
            return jsonify({"error": "Unable to generate schedule due to invalid employee data"}), 400
        
        end_time = datetime.now()
        process_time = (end_time - start_time).total_seconds()
        logger.info(f"Schedule generated in {process_time:.2f} seconds")
        
        return jsonify({
            'schedule': schedule,
            'month': month,
            'year': year,
            'month_name': calendar.month_name[month],
            'process_time': process_time
        })
    except Exception as e:
        logger.error(f"Error generating schedule: {str(e)}", exc_info=True)
        return jsonify({"error": str(e)}), 500

@app.route('/export', methods=['POST'])
def export():
    try:
        start_time = datetime.now()
        data = request.get_json()
        
        if not data:
            logger.warning("No JSON data received for export")
            return jsonify({"error": "No data provided"}), 400
            
        schedule = data.get('schedule', {})
        month = data.get('month')
        year = data.get('year')
        month_name = data.get('month_name')
        
        if not schedule or not month or not year or not month_name:
            logger.warning("Incomplete schedule data received for export")
            return jsonify({"error": "Incomplete schedule data provided"}), 400
        
        logger.info(f"Exporting schedule for {month}/{year} with {len(schedule)} employees")
        
        # Create a BytesIO object to store the Excel file
        output = BytesIO()
        
        wb = Workbook()
        ws = wb.active
        
        # Styles
        header_fill = PatternFill(start_color="1e3a8a", end_color="1e3a8a", fill_type="solid")
        subheader_fill = PatternFill(start_color="3b82f6", end_color="3b82f6", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        title_font = Font(color="FFFFFF", bold=True, size=16)
        border = Border(
            left=Side(style='thin', color="000000"),
            right=Side(style='thin', color="000000"),
            top=Side(style='thin', color="000000"),
            bottom=Side(style='thin', color="000000")
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Calculate last column letter
        num_days = calendar.monthrange(int(year), int(month))[1]
        last_col = get_column_letter(num_days + 3)
        
        # Set column widths first
        ws.column_dimensions['A'].width = 5  # S.R
        ws.column_dimensions['B'].width = 20  # SUPERVISOR
        ws.column_dimensions['C'].width = 10  # CODE NO.
        
        # Set fixed width for day columns
        for day in range(1, num_days + 1):
            col = get_column_letter(day + 3)
            ws.column_dimensions[col].width = 5
        
        # Main Header
        ws.merge_cells(f'A1:{last_col}1')
        cell = ws['A1']
        cell.value = 'BAGASSE YARD SHIFT SCHEDULE'
        cell.font = title_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        
        ws.merge_cells(f'A2:{last_col}2')
        cell = ws['A2']
        cell.value = f"{month_name.upper()} {year}"
        cell.font = title_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        
        # Column Headers
        ws['A3'] = 'S.R'
        ws['B3'] = 'SUPERVISOR'
        ws['C3'] = 'CODE NO.'
        
        # Day headers (1, 2, 3...)
        for day in range(1, num_days + 1):
            col = get_column_letter(day + 3)
            ws[f'{col}3'] = str(day)
            ws[f'{col}3'].font = header_font
            ws[f'{col}3'].fill = subheader_fill
            ws[f'{col}3'].alignment = center_alignment
            ws[f'{col}3'].border = border
            
            # Get the weekday name (Monday, Tuesday...)
            weekday_name = DutyScheduler.get_day_name(int(year), int(month), day)
            ws[f'{col}4'] = weekday_name[:3].upper()  # Using first 3 letters (MON, TUE...)
            ws[f'{col}4'].font = header_font
            ws[f'{col}4'].fill = subheader_fill
            ws[f'{col}4'].alignment = center_alignment
            ws[f'{col}4'].border = border
        
        # Apply styles to header row
        for col_letter in ['A', 'B', 'C']:
            ws[f'{col_letter}3'].font = header_font
            ws[f'{col_letter}3'].fill = subheader_fill
            ws[f'{col_letter}3'].alignment = center_alignment
            ws[f'{col_letter}3'].border = border
            
            # Weekday name row style
            ws[f'{col_letter}4'].font = header_font
            ws[f'{col_letter}4'].fill = subheader_fill
            ws[f'{col_letter}4'].alignment = center_alignment
            ws[f'{col_letter}4'].border = border
        
        # Fill data rows
        row_index = 5
        sr_no = 1
        
        for name, data in schedule.items():
            ws[f'A{row_index}'] = sr_no
            ws[f'B{row_index}'] = name
            ws[f'C{row_index}'] = data['code']
            
            shifts = data.get('shifts', [])
            for day, shift in enumerate(shifts, start=1):
                col = get_column_letter(day + 3)
                ws[f'{col}{row_index}'] = shift
                ws[f'{col}{row_index}'].alignment = center_alignment
                ws[f'{col}{row_index}'].border = border
            
            # Apply styles to the employee row
            for col_letter in ['A', 'B', 'C']:
                ws[f'{col_letter}{row_index}'].alignment = center_alignment
                ws[f'{col_letter}{row_index}'].border = border
            
            row_index += 1
            sr_no += 1
            
        # Add legend for shift codes
        legend_row = row_index + 2
        ws[f'A{legend_row}'] = 'Shift Legend:'
        ws[f'A{legend_row}'].font = Font(bold=True)
        
        legend_items = [
            ('A', 'Morning Shift (06:00-14:00)'),
            ('B', 'Afternoon Shift (14:00-22:00)'),
            ('C', 'Night Shift (22:00-06:00)'),
            ('G', 'General Shift (09:00-17:00)'),
            ('R', 'Rest Day')
        ]
        
        for i, (code, description) in enumerate(legend_items):
            ws[f'A{legend_row + i + 1}'] = f'{code} - {description}'
        
        # Save the Excel file
        wb.save(output)
        output.seek(0)
        
        end_time = datetime.now()
        process_time = (end_time - start_time).total_seconds()
        logger.info(f"Excel export completed in {process_time:.2f} seconds")
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'duty_schedule_{month_name}_{year}.xlsx'
        )
        
    except Exception as e:
        logger.error(f"Error exporting to Excel: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/debug_data', methods=['POST'])
def debug_data():
    """
    Debug endpoint to check the structure of received employee data
    """
    try:
        data = request.get_json()
        if not data:
            return jsonify({"error": "No data provided"}), 400
            
        employees_data = data.get('employees', [])
        if not employees_data:
            return jsonify({"error": "No employee data provided"}), 400
        
        result = {
            'data_type': str(type(employees_data)),
            'length': len(employees_data),
            'first_item_type': str(type(employees_data[0])) if employees_data else None,
            'first_item': employees_data[0] if employees_data else None,
            'all_items': employees_data
        }
        
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/export_pdf', methods=['POST'])
def export_pdf():
    """
    Export the schedule as a PDF file with attractive styling
    """
    try:
        start_time = datetime.now()
        data = request.get_json()
        
        if not data:
            logger.warning("No JSON data received for PDF export")
            return jsonify({"error": "No data provided"}), 400
            
        schedule = data.get('schedule', {})
        month = data.get('month')
        year = data.get('year')
        month_name = data.get('month_name')
        
        if not schedule or not month or not year or not month_name:
            logger.warning("Incomplete schedule data received for PDF export")
            return jsonify({"error": "Incomplete schedule data provided"}), 400
        
        logger.info(f"Exporting PDF schedule for {month}/{year} with {len(schedule)} employees")
        
        # Get the post name from the first employee (all employees will have same post)
        first_employee = next(iter(schedule.values()))
        post_name = first_employee.get('post', 'SUPERVISOR')
        
        # Create a PDF file in memory
        pdf_output = BytesIO()
        
        # Create the PDF document with adjusted margins
        doc = SimpleDocTemplate(
            pdf_output,
            pagesize=landscape(A4),
            rightMargin=10,
            leftMargin=10,
            topMargin=20,
            bottomMargin=20
        )
        
        # Get available page width and height
        page_width = landscape(A4)[0] - doc.rightMargin - doc.leftMargin
        page_height = landscape(A4)[1] - doc.topMargin - doc.bottomMargin
        
        elements = []
        
        # Add attractive title
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontSize=14,
            spaceAfter=20,
            alignment=1
        )
        
        title = Paragraph(
            f"<b>BAGASSE YARD SHIFT SCHEDULE - {post_name}</b><br/>{month_name.upper()} {year}",
            title_style
        )
        elements.append(title)
        
        # Prepare table data
        num_days = calendar.monthrange(int(year), int(month))[1]
        table_data = []
        
        # Headers row
        headers = ['SR', 'NAME', 'CD']  # Shortened headers
        headers.extend([str(day) for day in range(1, num_days + 1)])
        table_data.append(headers)
        
        # Day names row
        day_names = ['', '', '']  # Empty cells for SR, NAME, CD
        day_names.extend([DutyScheduler.get_day_name(int(year), int(month), day)[:3] 
                         for day in range(1, num_days + 1)])
        table_data.append(day_names)

        # Create post name cell with custom style
        post_style = ParagraphStyle(
            'PostStyle',
            parent=styles['Normal'],
            fontSize=10,  # Larger font size for post name
            textColor=colors.white,
            alignment=1,
            fontName='Helvetica-Bold'
        )
        post_cell = Paragraph(f"<b>{post_name}</b>", post_style)
        table_data[1][1] = post_cell  # Replace the second cell in day names row with styled post name

        # Employee data rows
        sr_no = 1
        for name, emp_data in schedule.items():
            row = [sr_no, name, emp_data['code']]
            row.extend(emp_data.get('shifts', []))
            table_data.append(row)
            sr_no += 1
        
        # Calculate optimal column widths
        name_col_width = page_width * 0.15  # 15% for name
        sr_col_width = page_width * 0.04   # 4% for serial number
        code_col_width = page_width * 0.04  # 4% for code
        remaining_width = page_width - (name_col_width + sr_col_width + code_col_width)
        day_width = remaining_width / num_days
        
        col_widths = [sr_col_width, name_col_width, code_col_width]
        col_widths.extend([day_width] * num_days)
        
        # Create table with optimized settings
        table = Table(table_data, colWidths=col_widths, rowHeights=[20]*len(table_data))
        
        # Style the table with attractive formatting
        table_style = TableStyle([
            # Headers
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 7),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            
            # Day names row with darker background for post name
            ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#1e3a8a')),  # Darker blue background
            ('TEXTCOLOR', (0, 1), (-1, 1), colors.white),
            ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 1), (-1, 1), 6),
            ('ALIGN', (0, 1), (-1, 1), 'CENTER'),
            
            # Data rows
            ('FONTNAME', (0, 2), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 2), (-1, -1), 7),  # Slightly larger font for data
            ('ALIGN', (0, 2), (-1, -1), 'CENTER'),  # Center all data
            
            # Grid styling
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),  # Lighter grid color
            ('LINEABOVE', (0, 0), (-1, 0), 1, colors.HexColor('#1e3a8a')),  # Thicker top border
            ('LINEBELOW', (0, 1), (-1, 1), 1, colors.HexColor('#3b82f6')),  # Thicker header bottom border
            
            # Cell alignment and padding
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 1),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
            ('LEFTPADDING', (0, 0), (-1, -1), 2),
            ('RIGHTPADDING', (0, 0), (-1, -1), 2),
            
            # Zebra striping for better readability
            ('ROWBACKGROUNDS', (0, 2), (-1, -1), [colors.HexColor('#f8fafc'), colors.white]),
        ])
        
        # Add shift-specific styles
        for row in range(2, len(table_data)):
            for col in range(3, len(table_data[row])):
                shift = table_data[row][col]
                if shift == 'A':
                    table_style.add('BACKGROUND', (col, row), (col, row), colors.HexColor('#dbeafe'))
                    table_style.add('TEXTCOLOR', (col, row), (col, row), colors.HexColor('#1e40af'))
                elif shift == 'B':
                    table_style.add('BACKGROUND', (col, row), (col, row), colors.HexColor('#ede9fe'))
                    table_style.add('TEXTCOLOR', (col, row), (col, row), colors.HexColor('#5b21b6'))
                elif shift == 'C':
                    table_style.add('BACKGROUND', (col, row), (col, row), colors.HexColor('#fff7ed'))
                    table_style.add('TEXTCOLOR', (col, row), (col, row), colors.HexColor('#c2410c'))
                elif shift == 'G':
                    table_style.add('BACKGROUND', (col, row), (col, row), colors.HexColor('#ccfbf1'))
                    table_style.add('TEXTCOLOR', (col, row), (col, row), colors.HexColor('#0f766e'))
                elif shift == 'R':
                    table_style.add('BACKGROUND', (col, row), (col, row), colors.HexColor('#f1f5f9'))
                    table_style.add('TEXTCOLOR', (col, row), (col, row), colors.HexColor('#334155'))
                table_style.add('FONTNAME', (col, row), (col, row), 'Helvetica-Bold')
        
        table.setStyle(table_style)
        elements.append(table)
        
        # Add legend
        elements.append(Spacer(1, 10))
        
        # Create legend table with simple text
        legend_data = [
            ['Shift Legend:', '', '', '', ''],
            [
                'A - Morning (06:00-14:00)',
                'B - Afternoon (14:00-22:00)',
                'C - Night (22:00-06:00)',
                'G - General (09:00-17:00)',
                'R - Rest Day'
            ]
        ]
        
        legend_table = Table(legend_data, colWidths=[page_width/5]*5, rowHeights=[12, 15])
        
        # Style the legend with colors directly in the table style
        legend_style = TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            # Add matching colors and backgrounds for each shift in the legend
            ('BACKGROUND', (0, 1), (0, 1), colors.HexColor('#dbeafe')),
            ('TEXTCOLOR', (0, 1), (0, 1), colors.HexColor('#1e40af')),
            ('BACKGROUND', (1, 1), (1, 1), colors.HexColor('#ede9fe')),
            ('TEXTCOLOR', (1, 1), (1, 1), colors.HexColor('#5b21b6')),
            ('BACKGROUND', (2, 1), (2, 1), colors.HexColor('#fff7ed')),
            ('TEXTCOLOR', (2, 1), (2, 1), colors.HexColor('#c2410c')),
            ('BACKGROUND', (3, 1), (3, 1), colors.HexColor('#ccfbf1')),
            ('TEXTCOLOR', (3, 1), (3, 1), colors.HexColor('#0f766e')),
            ('BACKGROUND', (4, 1), (4, 1), colors.HexColor('#f1f5f9')),
            ('TEXTCOLOR', (4, 1), (4, 1), colors.HexColor('#334155')),
        ])
        
        legend_table.setStyle(legend_style)
        elements.append(legend_table)
        
        # Build PDF
        doc.build(elements)
        
        # Prepare for download
        pdf_output.seek(0)
        
        end_time = datetime.now()
        process_time = (end_time - start_time).total_seconds()
        logger.info(f"PDF export completed in {process_time:.2f} seconds")
        
        return send_file(
            pdf_output,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'duty_schedule_{month_name}_{year}.pdf'
        )
        
    except Exception as e:
        logger.error(f"Error exporting to PDF: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.errorhandler(404)
def page_not_found(e):
    return jsonify({"error": "Resource not found"}), 404

@app.errorhandler(500)
def internal_server_error(e):
    return jsonify({"error": "Internal server error"}), 500

if __name__ == '__main__':
    # Get port from environment variable or use 5000 as default
    port = int(os.environ.get('PORT', 5000))
    # Set debug mode based on environment variable
    debug = os.environ.get('FLASK_DEBUG', 'False').lower() == 'true'
    app.run(host='0.0.0.0', port=port, debug=debug) 