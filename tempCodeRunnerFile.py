import os
from flask import Flask, render_template, request, send_file, jsonify
from datetime import datetime, timedelta
import calendar
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, Color
from openpyxl.utils import get_column_letter
from functools import lru_cache
import logging
from werkzeug.middleware.profiler import ProfilerMiddleware
from io import BytesIO

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

app = Flask(__name__)

# Enable profiling conditionally
if os.environ.get('ENABLE_PROFILING') == '1':
    app.config['PROFILE'] = True
    app.wsgi_app = ProfilerMiddleware(app.wsgi_app, restrictions=[30])

class DutyScheduler:
    def __init__(self):
        self.shifts = {
            'A': '06:00-14:00',
            'B': '14:00-22:00',
            'C': '22:00-06:00',
            'G': 'General',
            'R': 'Rest'
        }
        self.shift_rotation = {'A': 'C', 'C': 'B', 'B': 'A'}
        # Color mappings for the Excel output
        self.shift_colors = {
            'A': '3b82f6',  # Blue
            'B': '8b5cf6',  # Purple
            'C': 'f97316',  # Orange
            'G': '14b8a6',  # Teal
            'R': '64748b'   # Gray
        }
        
    @staticmethod
    def get_day_name(year, month, day):
        return calendar.day_abbr[calendar.weekday(year, month, day)].upper()

    def get_next_shift(self, current_shift):
        return self.shift_rotation.get(current_shift, current_shift)

    def generate_schedule(self, employees_data, year, month):
        """
        Generate the duty schedule
        
        Args:
            employees_data: List or tuple containing employee data
            year: The year to generate the schedule for
            month: The month to generate the schedule for
            
        Returns:
            A dictionary with the schedule for each employee
        """
        # 1. Debug input information
        logger.info(f"Input employees_data type: {type(employees_data)}")
        logger.info(f"Example of employees_data: {str(employees_data)[:200]}")  # Limit output size
        
        # 2. Ensure we're working with a list
        if not isinstance(employees_data, list):
            employees_data = list(employees_data) if isinstance(employees_data, tuple) else [employees_data]
        
        # 3. Create a clean list of employee dictionaries
        cleaned_employees = []
        
        for emp in employees_data:
            # If we have a tuple, convert it to a dict
            if isinstance(emp, tuple):
                emp_dict = {}
                try:
                    # Handle tuple of key-value tuples
                    for item in emp:
                        if isinstance(item, tuple) and len(item) >= 2:
                            key, value = item[0], item[1]
                            emp_dict[key] = value
                except Exception as e:
                    logger.error(f"Error converting tuple to dict: {str(e)}, tuple: {emp}")
                    continue
            elif isinstance(emp, dict):
                # Already a dict
                emp_dict = emp
            else:
                # Unknown format
                logger.error(f"Unsupported employee data type: {type(emp)}, value: {emp}")
                continue
            
            # Ensure all required keys are present
            required_keys = ['name', 'code', 'post', 'start_shift', 'rest_day']
            if not all(key in emp_dict for key in required_keys):
                missing = [key for key in required_keys if key not in emp_dict]
                logger.error(f"Missing required keys {missing} in employee data: {emp_dict}")
                continue
            
            # Convert rest_day to int if it's not already
            try:
                emp_dict['rest_day'] = int(emp_dict['rest_day'])
            except (ValueError, TypeError):
                logger.error(f"Invalid rest_day value: {emp_dict.get('rest_day')}")
                continue
                
            cleaned_employees.append(emp_dict)
        
        # If we have no valid employees, return an empty schedule
        if not cleaned_employees:
            logger.error("No valid employees after data cleaning")
            return {}
            
        # 4. Generate the schedule
        logger.info(f"Generating schedule with {len(cleaned_employees)} valid employees")
        num_days = calendar.monthrange(year, month)[1]
        schedule = {}
        
        # Pre-calculate weekdays for the entire month
        weekdays = [calendar.weekday(year, month, day) for day in range(1, num_days + 1)]
        
        for emp in cleaned_employees:
            schedule[emp['name']] = {
                'code': emp['code'],
                'post': emp['post'],
                'shifts': []
            }
            
            current_shift = emp['start_shift']
            # Convert Sunday from 0 to 6
            rest_day = 6 if emp['rest_day'] == 0 else emp['rest_day'] - 1
            was_rest_day = False
            
            for day_index, day in enumerate(range(1, num_days + 1)):
                # Check if it's a rest day using pre-calculated weekdays
                if weekdays[day_index] == rest_day:
                    schedule[emp['name']]['shifts'].append('R')
                    was_rest_day = True
                else:
                    if was_rest_day and current_shift != 'G':
                        current_shift = self.get_next_shift(current_shift)
                        was_rest_day = False
                    schedule[emp['name']]['shifts'].append(current_shift)
        
        return schedule

@app.route('/')
def index():
    try:
        logger.info("Rendering index page")
        return render_template('index.html')
    except Exception as e:
        logger.error(f"Error rendering index page: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/generate', methods=['POST'])
def generate():
    try:
        start_time = datetime.now()
        data = request.get_json()
        
        if not data:
            logger.warning("No JSON data received")
            return jsonify({"error": "No data provided"}), 400
            
        year = int(data.get('year', datetime.now().year))
        month = int(data.get('month', datetime.now().month))
        employees_data = data.get('employees', [])
        
        # Add detailed logging for debugging
        logger.info(f"Raw data received: {str(data)[:200]}")
        logger.info(f"Employees data type: {type(employees_data)}")
        
        if not employees_data:
            logger.warning("No employee data received")
            return jsonify({"error": "No employee data provided"}), 400
        
        logger.info(f"Generating schedule for {month}/{year} with {len(employees_data)} employees")
        
        # Try to see what's in the first employee
        if employees_data and len(employees_data) > 0:
            first_emp = employees_data[0]
            logger.info(f"First employee type: {type(first_emp)}")
            logger.info(f"First employee data: {first_emp}")
            
            # If it's a dictionary, check the keys
            if isinstance(first_emp, dict):
                logger.info(f"First employee keys: {first_emp.keys()}")
        
        # Generate the schedule
        scheduler = DutyScheduler()
        schedule = scheduler.generate_schedule(employees_data, year, month)
        
        # Check if we got an empty schedule (indicates error)
        if not schedule:
            logger.error("Generated schedule is empty, likely due to data errors")
            return jsonify({"error": "Unable to generate schedule due to invalid employee data"}), 400
        
        end_time = datetime.now()
        process_time = (end_time - start_time).total_seconds()
        logger.info(f"Schedule generated in {process_time:.2f} seconds")
        
        return jsonify({
            'schedule': schedule,
            'month': month,
            'year': year,
            'month_name': calendar.month_name[month],
            'process_time': process_time
        })
    except Exception as e:
        logger.error(f"Error generating schedule: {str(e)}", exc_info=True)
        return jsonify({"error": str(e)}), 500

@app.route('/export', methods=['POST'])
def export():
    try:
        start_time = datetime.now()
        data = request.get_json()
        
        if not data:
            logger.warning("No JSON data received for export")
            return jsonify({"error": "No data provided"}), 400
            
        schedule = data.get('schedule', {})
        month = data.get('month')
        year = data.get('year')
        month_name = data.get('month_name')
        
        if not schedule or not month or not year or not month_name:
            logger.warning("Incomplete schedule data received for export")
            return jsonify({"error": "Incomplete schedule data provided"}), 400
        
        logger.info(f"Exporting schedule for {month}/{year} with {len(schedule)} employees")
        
        # Create a BytesIO object to store the Excel file
        output = BytesIO()
        
        wb = Workbook()
        ws = wb.active
        
        # Styles
        header_fill = PatternFill(start_color="1e3a8a", end_color="1e3a8a", fill_type="solid")
        subheader_fill = PatternFill(start_color="3b82f6", end_color="3b82f6", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        title_font = Font(color="FFFFFF", bold=True, size=16)
        border = Border(
            left=Side(style='thin', color="000000"),
            right=Side(style='thin', color="000000"),
            top=Side(style='thin', color="000000"),
            bottom=Side(style='thin', color="000000")
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Calculate last column letter
        num_days = calendar.monthrange(int(year), int(month))[1]
        last_col = get_column_letter(num_days + 3)
        
        # Main Header
        ws.merge_cells(f'A1:{last_col}1')
        ws['A1'] = 'BAGASSE YARD SHIFT SCHEDULE'
        ws['A1'].font = title_font
        ws['A1'].fill = header_fill
        ws['A1'].alignment = center_alignment
        
        ws.merge_cells(f'A2:{last_col}2')
        ws['A2'] = f"{month_name.upper()} {year}"
        ws['A2'].font = title_font
        ws['A2'].fill = header_fill
        ws['A2'].alignment = center_alignment
        
        # Column Headers
        ws['A3'] = 'S.R'
        ws['B3'] = 'SUPERVISOR'
        ws['C3'] = 'CODE NO.'
        
        # Day headers (1, 2, 3...)
        for day in range(1, num_days + 1):
            col = get_column_letter(day + 3)
            ws[f'{col}3'] = str(day)
            ws[f'{col}3'].font = header_font
            ws[f'{col}3'].fill = subheader_fill
            ws[f'{col}3'].alignment = center_alignment
            ws[f'{col}3'].border = border
            
            # Get the weekday name (Monday, Tuesday...)
            weekday_name = DutyScheduler.get_day_name(int(year), int(month), day)
            ws[f'{col}4'] = weekday_name[:3].upper()  # Using first 3 letters (MON, TUE...)
            ws[f'{col}4'].font = header_font
            ws[f'{col}4'].fill = subheader_fill
            ws[f'{col}4'].alignment = center_alignment
            ws[f'{col}4'].border = border
        
        # Apply styles to header row
        for col_letter in ['A', 'B', 'C']:
            ws[f'{col_letter}3'].font = header_font
            ws[f'{col_letter}3'].fill = subheader_fill
            ws[f'{col_letter}3'].alignment = center_alignment
            ws[f'{col_letter}3'].border = border
            
            # Weekday name row style
            ws[f'{col_letter}4'].font = header_font
            ws[f'{col_letter}4'].fill = subheader_fill
            ws[f'{col_letter}4'].alignment = center_alignment
            ws[f'{col_letter}4'].border = border
        
        # Fill data rows
        row_index = 5
        sr_no = 1
        
        for name, data in schedule.items():
            ws[f'A{row_index}'] = sr_no
            ws[f'B{row_index}'] = name
            ws[f'C{row_index}'] = data['code']
            
            shifts = data.get('shifts', [])
            for day, shift in enumerate(shifts, start=1):
                col = get_column_letter(day + 3)
                ws[f'{col}{row_index}'] = shift
                ws[f'{col}{row_index}'].alignment = center_alignment
                ws[f'{col}{row_index}'].border = border
            
            # Apply styles to the employee row
            for col_letter in ['A', 'B', 'C']:
                ws[f'{col_letter}{row_index}'].alignment = center_alignment
                ws[f'{col_letter}{row_index}'].border = border
            
            row_index += 1
            sr_no += 1
            
        # Add legend for shift codes
        legend_row = row_index + 2
        ws[f'A{legend_row}'] = 'Shift Legend:'
        ws[f'A{legend_row}'].font = Font(bold=True)
        
        legend_items = [
            ('A', 'Morning Shift (06:00-14:00)'),
            ('B', 'Afternoon Shift (14:00-22:00)'),
            ('C', 'Night Shift (22:00-06:00)'),
            ('G', 'General Shift (09:00-17:00)'),
            ('R', 'Rest Day')
        ]
        
        for i, (code, description) in enumerate(legend_items):
            ws[f'A{legend_row + i + 1}'] = f'{code} - {description}'
        
        # Auto-adjust column width
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        
        # Set fixed width for day columns (to make them all consistent)
        for day in range(1, num_days + 1):
            col = get_column_letter(day + 3)
            ws.column_dimensions[col].width = 5
        
        # Save the Excel file
        wb.save(output)
        output.seek(0)
        
        end_time = datetime.now()
        process_time = (end_time - start_time).total_seconds()
        logger.info(f"Excel export completed in {process_time:.2f} seconds")
        
        filename = f'duty_schedule_{month_name}_{year}.xlsx'
        
        response = send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
        # Add proper headers
        response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.headers['Access-Control-Expose-Headers'] = 'Content-Disposition'
        
        return response
        
    except Exception as e:
        logger.error(f"Error exporting to Excel: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/debug_data', methods=['POST'])
def debug_data():
    """
    Debug endpoint to check the structure of received employee data
    """
    try:
        data = request.get_json()
        if not data:
            return jsonify({"error": "No data provided"}), 400
            
        employees_data = data.get('employees', [])
        if not employees_data:
            return jsonify({"error": "No employee data provided"}), 400
        
        result = {
            'data_type': str(type(employees_data)),
            'length': len(employees_data),
            'first_item_type': str(type(employees_data[0])) if employees_data else None,
            'first_item': employees_data[0] if employees_data else None,
            'all_items': employees_data
        }
        
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/export_pdf', methods=['POST'])
def export_pdf():
    """
    Export the schedule as a PDF file
    """
    try:
        start_time = datetime.now()
        data = request.get_json()
        
        if not data:
            logger.warning("No JSON data received for PDF export")
            return jsonify({"error": "No data provided"}), 400
            
        schedule = data.get('schedule', {})
        month = data.get('month')
        year = data.get('year')
        month_name = data.get('month_name')
        
        if not schedule or not month or not year or not month_name:
            logger.warning("Incomplete schedule data received for PDF export")
            return jsonify({"error": "Incomplete schedule data provided"}), 400
        
        logger.info(f"Exporting PDF schedule for {month}/{year} with {len(schedule)} employees")
        
        # First create an Excel file in memory
        output = BytesIO()
        
        wb = Workbook()
        ws = wb.active
        
        # Styles
        header_fill = PatternFill(start_color="1e3a8a", end_color="1e3a8a", fill_type="solid")
        subheader_fill = PatternFill(start_color="3b82f6", end_color="3b82f6", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        title_font = Font(color="FFFFFF", bold=True, size=16)
        border = Border(
            left=Side(style='thin', color="000000"),
            right=Side(style='thin', color="000000"),
            top=Side(style='thin', color="000000"),
            bottom=Side(style='thin', color="000000")
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Calculate last column letter
        num_days = calendar.monthrange(int(year), int(month))[1]
        last_col = get_column_letter(num_days + 3)
        
        # Main Header
        ws.merge_cells(f'A1:{last_col}1')
        ws['A1'] = 'BAGASSE YARD SHIFT SCHEDULE'
        ws['A1'].font = title_font
        ws['A1'].fill = header_fill
        ws['A1'].alignment = center_alignment
        
        ws.merge_cells(f'A2:{last_col}2')
        ws['A2'] = f"{month_name.upper()} {year}"
        ws['A2'].font = title_font
        ws['A2'].fill = header_fill
        ws['A2'].alignment = center_alignment
        
        # Column Headers
        ws['A3'] = 'S.R'
        ws['B3'] = 'SUPERVISOR'
        ws['C3'] = 'CODE NO.'
        
        # Day headers (1, 2, 3...)
        for day in range(1, num_days + 1):
            col = get_column_letter(day + 3)
            ws[f'{col}3'] = str(day)
            ws[f'{col}3'].font = header_font
            ws[f'{col}3'].fill = subheader_fill
            ws[f'{col}3'].alignment = center_alignment
            ws[f'{col}3'].border = border
            
            # Get the weekday name (Monday, Tuesday...)
            weekday_name = DutyScheduler.get_day_name(int(year), int(month), day)
            ws[f'{col}4'] = weekday_name[:3].upper()  # Using first 3 letters (MON, TUE...)
            ws[f'{col}4'].font = header_font
            ws[f'{col}4'].fill = subheader_fill
            ws[f'{col}4'].alignment = center_alignment
            ws[f'{col}4'].border = border
        
        # Apply styles to header row
        for col_letter in ['A', 'B', 'C']:
            ws[f'{col_letter}3'].font = header_font
            ws[f'{col_letter}3'].fill = subheader_fill
            ws[f'{col_letter}3'].alignment = center_alignment
            ws[f'{col_letter}3'].border = border
            
            # Weekday name row style
            ws[f'{col_letter}4'].font = header_font
            ws[f'{col_letter}4'].fill = subheader_fill
            ws[f'{col_letter}4'].alignment = center_alignment
            ws[f'{col_letter}4'].border = border
        
        # Fill data rows
        row_index = 5
        sr_no = 1
        
        for name, data in schedule.items():
            ws[f'A{row_index}'] = sr_no
            ws[f'B{row_index}'] = name
            ws[f'C{row_index}'] = data['code']
            
            shifts = data.get('shifts', [])
            for day, shift in enumerate(shifts, start=1):
                col = get_column_letter(day + 3)
                ws[f'{col}{row_index}'] = shift
                ws[f'{col}{row_index}'].alignment = center_alignment
                ws[f'{col}{row_index}'].border = border
            
            # Apply styles to the employee row
            for col_letter in ['A', 'B', 'C']:
                ws[f'{col_letter}{row_index}'].alignment = center_alignment
                ws[f'{col_letter}{row_index}'].border = border
            
            row_index += 1
            sr_no += 1
            
        # Add legend for shift codes
        legend_row = row_index + 2
        ws[f'A{legend_row}'] = 'Shift Legend:'
        ws[f'A{legend_row}'].font = Font(bold=True)
        
        legend_items = [
            ('A', 'Morning Shift (06:00-14:00)'),
            ('B', 'Afternoon Shift (14:00-22:00)'),
            ('C', 'Night Shift (22:00-06:00)'),
            ('G', 'General Shift (09:00-17:00)'),
            ('R', 'Rest Day')
        ]
        
        for i, (code, description) in enumerate(legend_items):
            ws[f'A{legend_row + i + 1}'] = f'{code} - {description}'
        
        # Auto-adjust column width
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        
        # Set fixed width for day columns (to make them all consistent)
        for day in range(1, num_days + 1):
            col = get_column_letter(day + 3)
            ws.column_dimensions[col].width = 5
        
        # Save the Excel file to memory
        wb.save(output)
        output.seek(0)
        
        # Convert the Excel to PDF
        from openpyxl import load_workbook
        from reportlab.lib.pagesizes import landscape, A4
        from reportlab.pdfgen import canvas
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
        from reportlab.lib import colors
        from reportlab.platypus import Paragraph
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        
        # Load workbook from the BytesIO object
        wb = load_workbook(output)
        ws = wb.active
        
        # Create a PDF file in memory
        pdf_output = BytesIO()
        doc = SimpleDocTemplate(pdf_output, pagesize=landscape(A4))
        
        # Extract data from Excel
        data = []
        
        # Get headers first
        headers = []
        headers.append("S.R")
        headers.append("SUPERVISOR")
        headers.append("CODE NO.")
        for day in range(1, num_days + 1):
            headers.append(str(day))
        data.append(headers)
        
        # Get day names
        day_names = ["", "", ""]
        for day in range(1, num_days + 1):
            day_names.append(DutyScheduler.get_day_name(int(year), int(month), day)[:3].upper())
        data.append(day_names)
        
        # Get employee data
        for emp_idx in range(5, row_index):
            row = []
            row.append(ws[f'A{emp_idx}'].value)  # S.R
            row.append(ws[f'B{emp_idx}'].value)  # Name
            row.append(ws[f'C{emp_idx}'].value)  # Code
            
            for day in range(1, num_days + 1):
                col = get_column_letter(day + 3)
                row.append(ws[f'{col}{emp_idx}'].value)  # Shift for this day
            
            data.append(row)
        
        # Create a table with the data
        table = Table(data)
        
        # Style the table
        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
            ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (-1, 1), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 2), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ])
        table.setStyle(style)
        
        # Create a title
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'Title', 
            parent=styles['Heading1'], 
            alignment=1,  # Center
            textColor=colors.HexColor('#1e3a8a'),
            spaceAfter=20
        )
        
        title = Paragraph(f"BAGASSE YARD SHIFT SCHEDULE - {month_name.upper()} {year}", title_style)
        
        # Create a legend
        legend_style = ParagraphStyle(
            'Legend', 
            parent=styles['Normal'], 
            alignment=0,  # Left
            spaceAfter=5
        )
        
        legend_items_pdf = [
            Paragraph("Shift Legend:", ParagraphStyle('LegendTitle', parent=legend_style, fontName='Helvetica-Bold')),
            Paragraph("A - Morning Shift (06:00-14:00)", legend_style),
            Paragraph("B - Afternoon Shift (14:00-22:00)", legend_style),
            Paragraph("C - Night Shift (22:00-06:00)", legend_style),
            Paragraph("G - General Shift (09:00-17:00)", legend_style),
            Paragraph("R - Rest Day", legend_style)
        ]
        
        # Build the PDF
        elements = [title, table]
        elements.extend(legend_items_pdf)
        doc.build(elements)
        
        # Prepare the PDF for download
        pdf_output.seek(0)
        
        end_time = datetime.now()
        process_time = (end_time - start_time).total_seconds()
        logger.info(f"PDF export completed in {process_time:.2f} seconds")
        
        return send_file(
            pdf_output,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'duty_schedule_{month_name}_{year}.pdf'
        )
        
    except Exception as e:
        logger.error(f"Error exporting to PDF: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.errorhandler(404)
def page_not_found(e):
    return jsonify({"error": "Resource not found"}), 404

@app.errorhandler(500)
def internal_server_error(e):
    return jsonify({"error": "Internal server error"}), 500

if __name__ == '__main__':
    # Get port from environment variable or use 5000 as default
    port = int(os.environ.get('PORT', 5000))
    # Set debug mode based on environment variable
    debug = os.environ.get('FLASK_DEBUG', 'False').lower() == 'true'
    app.run(host='0.0.0.0', port=port, debug=debug) 